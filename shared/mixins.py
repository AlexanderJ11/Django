from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse
from django.utils import timezone

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


class StaffRequiredMixin:
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)
        return super().dispatch(request, *args, **kwargs)


def _get_field_value(obj, accessor):
    """Resolve a field value from an object using dot/double-underscore notation or callable."""
    if callable(accessor):
        return accessor(obj)
    parts = accessor.replace('__', '.').split('.')
    val = obj
    for part in parts:
        val = getattr(val, part, '')
        if val is None:
            return ''
        if callable(val) and not hasattr(val, 'all'):
            val = val()
    return val if val is not None else ''


class ExportMixin:
    """
    Mixin genérico para exportar el queryset filtrado a PDF o Excel.

    Uso en cualquier ListView:

        class ProductListView(LoginRequiredMixin, ExportMixin, ListView):
            export_filename = 'products'
            export_title    = 'Product List'
            export_fields   = [
                ('Name',     'name'),
                ('Brand',    'brand__name'),
                ('Price',    'unit_price'),
                ('Active',   lambda obj: 'Yes' if obj.is_active else 'No'),
            ]
    """

    export_filename = 'export'
    export_title    = 'Report'
    export_fields   = []          # [(header, accessor_or_callable), ...]

    # ------------------------------------------------------------------ #
    #  Intercept GET to check for ?export=pdf / ?export=excel             #
    # ------------------------------------------------------------------ #
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt == 'pdf':
            return self._export_pdf(request)
        if fmt == 'excel':
            return self._export_excel(request)
        return super().get(request, *args, **kwargs)

    # ------------------------------------------------------------------ #
    #  Build rows from the filtered queryset (no pagination)              #
    # ------------------------------------------------------------------ #
    def _get_export_data(self):
        headers = [h for h, _ in self.export_fields]
        rows = []
        for obj in self.get_queryset():
            rows.append([str(_get_field_value(obj, acc)) for _, acc in self.export_fields])
        return headers, rows

    # ------------------------------------------------------------------ #
    #  Excel export                                                        #
    # ------------------------------------------------------------------ #
    def _export_excel(self, request):
        headers, rows = self._get_export_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_title[:31]   # Excel sheet name max 31 chars

        # ── Title row ──
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=self.export_title)
        title_cell.font      = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill      = PatternFill('solid', fgColor='1E293B')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Date row ──
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2,   end_column=len(headers))
        date_str = timezone.now().strftime('%Y-%m-%d %H:%M')
        date_cell = ws.cell(row=2, column=1,
                            value=f'Generated: {date_str}  |  {len(rows)} records')
        date_cell.font      = Font(italic=True, size=9, color='64748B')
        date_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 16

        # ── Header row ──
        header_fill   = PatternFill('solid', fgColor='6366F1')
        header_font   = Font(bold=True, color='FFFFFF', size=10)
        header_border = Border(
            bottom=Side(style='medium', color='4F46E5')
        )
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = header_border
        ws.row_dimensions[3].height = 20

        # ── Data rows ──
        fill_even = PatternFill('solid', fgColor='F8FAFC')
        fill_odd  = PatternFill('solid', fgColor='FFFFFF')
        thin      = Side(style='thin', color='E2E8F0')
        data_border = Border(bottom=thin)

        for row_idx, row in enumerate(rows, start=4):
            fill = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = fill
                cell.border    = data_border
                cell.alignment = Alignment(vertical='center')
                cell.font      = Font(size=9)

        # ── Auto column widths ──
        for col_idx, header in enumerate(headers, start=1):
            col_values = [header] + [r[col_idx - 1] for r in rows]
            max_len = max((len(str(v)) for v in col_values), default=10)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 4, 40)

        # ── Response ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{self.export_filename}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        return response

    # ------------------------------------------------------------------ #
    #  PDF export                                                          #
    # ------------------------------------------------------------------ #
    def _export_pdf(self, request):
        headers, rows = self._get_export_data()
        buf = io.BytesIO()

        page_size = landscape(letter) if len(headers) > 5 else letter
        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.6 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ExportTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            'ExportSub',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=10,
        )

        elements = []

        # Title + subtitle
        elements.append(Paragraph(self.export_title, title_style))
        date_str = timezone.now().strftime('%Y-%m-%d %H:%M')
        elements.append(
            Paragraph(f'Generated: {date_str}  ·  {len(rows)} records', sub_style)
        )

        # Build table data
        usable_w = page_size[0] - inch
        col_w = usable_w / len(headers)
        table_data = [headers] + rows if rows else [headers, ['No data'] + [''] * (len(headers) - 1)]

        tbl = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
        tbl.setStyle(TableStyle([
            # Header
            ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 9),
            ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN',       (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING',   (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING',(0, 0), (-1, 0), 6),
            # Data rows
            ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',     (0, 1), (-1, -1), 8),
            ('ALIGN',        (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN',       (0, 1), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',   (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING',(0, 1), (-1, -1), 4),
            # Alternating rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.HexColor('#F8FAFC'), colors.white]),
            # Grid
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
            ('LINEBELOW',    (0, 0), (-1, 0), 1.2, colors.HexColor('#4F46E5')),
            # Rounded outer border feel
            ('BOX',          (0, 0), (-1, -1), 0.8, colors.HexColor('#CBD5E1')),
        ]))

        elements.append(tbl)

        # Page footer callback
        date_footer = date_str

        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#94A3B8'))
            w, h = page_size
            canvas.drawString(0.5 * inch, 0.3 * inch,
                              f'{self.export_title}  ·  {date_footer}')
            canvas.drawRightString(w - 0.5 * inch, 0.3 * inch,
                                   f'Page {doc.page}')
            canvas.restoreState()

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{self.export_filename}_{timezone.now().strftime("%Y%m%d")}.pdf"'
        )
        return response
